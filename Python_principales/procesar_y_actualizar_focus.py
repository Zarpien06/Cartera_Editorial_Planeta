import os
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
import json
import shutil
import traceback
import re
import time
import logging
import tempfile
from logging.handlers import RotatingFileHandler
from typing import Optional, Tuple, Dict, Any, List, Union
from pathlib import Path
from datetime import datetime, timedelta

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill

# Importar xlrd y xlwt para manejo de archivos .xls
try:
    import xlrd
    import xlwt
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

# Nombre del archivo de salida FOCUS procesado
OUTPUT_FILENAME = "FOCUS_ACTUALIZADO.xlsx"

# Directorios base del proyecto
BASE_DIR = Path(__file__).parent
SALIDAS_DIR = BASE_DIR / "salidas"
BACKUP_DIR = SALIDAS_DIR / "backup"

def encontrar_hoja_por_nombre(wb, nombre_buscar: str) -> str:
    nombre_buscar = str(nombre_buscar).strip().lower()
    for nombre_hoja in wb.sheetnames:
        if str(nombre_hoja).strip().lower() == nombre_buscar:
            return nombre_hoja
    raise ValueError(f"No se encontró la hoja: {nombre_buscar}")

# Marcador para identificar archivos generados por este script
GEN_MARKER_KEY = "_FOCUS_GENERATED_BY"
GEN_MARKER_VAL = "procesar_y_actualizar_focus_v1"

def workbook_has_generation_marker(wb: openpyxl.Workbook, ws: Worksheet) -> bool:
    """Detecta si el libro fue generado previamente por este script."""
    try:
        if str(ws["Z1"].value).strip() == GEN_MARKER_VAL:
            return True
    except Exception:
        pass
    try:
        subject = getattr(wb.properties, "subject", "") or ""
        return GEN_MARKER_VAL in str(subject)
    except Exception:
        return False

def set_workbook_generation_marker(wb: openpyxl.Workbook, ws: Worksheet) -> None:
    """Escribe el marcador en el libro y hoja activa para futuras detecciones."""
    try:
        ws["Z1"].value = GEN_MARKER_VAL
        ws.column_dimensions['Z'].hidden = True
    except Exception:
        pass
    try:
        props = wb.properties
        current_subject = props.subject or ""
        if GEN_MARKER_VAL not in current_subject:
            props.subject = (current_subject + " " + GEN_MARKER_VAL).strip()
    except Exception:
        pass

def limpiar_archivos_antiguos(directorio: Path, horas: int = 24) -> None:
    """Elimina archivos con antigüedad mayor a 'horas' en el directorio dado."""
    try:
        limite = datetime.now().timestamp() - horas * 3600
        if directorio.is_dir():
            for p in directorio.iterdir():
                try:
                    if p.is_file() and p.stat().st_mtime < limite:
                        p.unlink(missing_ok=True)
                except Exception:
                    pass
    except Exception:
        pass

# Limpieza inicial al cargar el módulo
limpiar_archivos_antiguos(SALIDAS_DIR, horas=24)
limpiar_archivos_antiguos(BACKUP_DIR, horas=24)

# Configurar el nivel de logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.addHandler(logging.StreamHandler())
logger.addHandler(RotatingFileHandler(BASE_DIR / 'focus_processor.log', maxBytes=1048576, backupCount=3, encoding='utf-8'))

from log_bridge import attach_python_logging, log_event
attach_python_logging('PY_FOCUS', logger)
logger.info('Iniciando procesador FOCUS')

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

CUENTAS_BALANCE = [
    # Cuentas explícitas requeridas por FORMATO DEUDA 1.pdf
    "Total cuenta objeto 43001",
    "Total cuenta objeto 43008",
    "Total cuenta objeto 43042",

    # Cuentas detalladas (rangos y cuentas específicas) - BALANCE
    "0080.43002.20", "0080.43002.21", "0080.43002.15",
    "0080.43002.28", "0080.43002.31", "0080.43002.63",

    # Cuentas adicionales de 43002
    "0080.43002.29", "0080.43002.57",
    "0080.43002.60", "0080.43002.62",
    "0080.43002.64", "0080.43002.65", "0080.43002.66", "0080.43002.69",

    # Cuentas 01010 (clientes/grupo planeta y subcuentas)
    "0080.01010.00010",  # CLIENTE GR.ED.PLANETA
    "0080.01010.00600",  # CLIENTE GR.PLANETA CHILE
    "0080.01010.10801", "0080.01010.10802", "0080.01010.10805",
    "0080.01010.10808", "0080.01010.10809", "0080.01010.10817",
    "0080.01010.10818", "0080.01010.10819",

    # Patrones más amplios para captura por prefijo
    "0080.01010",  # todas las cuentas de este rango
    "0080.43002",  # todas las cuentas de este rango
]


def obtener_mes_siguiente(mes_actual: str) -> str:
    """Dado un nombre de mes en español, devuelve el siguiente mes.
    Si no reconoce el mes, devuelve el mismo valor por seguridad."""
    try:
        idx = int(MESES.index(mes_actual))
        return MESES[(idx + 1) % 12]
    except Exception:
        # Intentar normalizar
        lower_map = {m.lower(): i for i, m in enumerate(MESES)}
        key = str(mes_actual).strip().lower()
        if key in lower_map:
            return MESES[(int(lower_map[key]) + 1) % 12]
        return mes_actual


def normalize_text(value: str) -> str:
    """Normaliza texto: minúsculas, preservando ñ y caracteres especiales del español."""
    try:
        import unicodedata
        s = str(value).lower()
        # Normalizar pero preservar ñ y otros caracteres especiales del español
        s = unicodedata.normalize('NFC', s)
        # Reemplazar separadores comunes pero preservar caracteres especiales
        return ''.join(ch if ch.isalnum() or ch.isspace() or ch in 'ñáéíóúü' else ' ' for ch in s)
    except Exception:
        return str(value).lower()


def to_float(valor) -> float:
    """Convierte cualquier valor a float de forma segura"""
    if pd.isna(valor) or valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    try:
        s = str(valor).strip().replace(',', '').replace(' ', '')
        return float(re.sub(r'[^\d.-]', '', s))
    except:
        return 0.0


def buscar_archivo(directorio: Path, patrones: list, extensiones=None, excluir: Optional[List[str]] = None) -> Optional[Path]:
    archivos = []
    excluir = excluir or ["backup", "old", "prev", "temp", "actualizado"]  # <-- agregar "actualizado"
    for archivo in directorio.iterdir():
        if archivo.is_file():
            nombre = archivo.name.lower()
            if any(patron.lower() in nombre for patron in patrones):
                if any(bad in nombre for bad in excluir):
                    continue
                if extensiones is None or any(nombre.endswith(ext.lower()) for ext in extensiones):
                    archivos.append(archivo)
    return max(archivos, key=lambda x: x.stat().st_mtime) if archivos else None

def leer_excel(ruta: Path, hoja=0):
    """
    Lee archivo Excel con manejo de errores y soporte para .xls y .xlsx
    
    Args:
        ruta: Ruta al archivo Excel
        hoja: Índice o nombre de la hoja a leer (por defecto: primera hoja)
        
    Returns:
        DataFrame de pandas con los datos de la hoja especificada
        
    Raises:
        ValueError: Si el formato del archivo no es soportado
        Exception: Si ocurre un error al leer el archivo
    """
    try:
        # Verificar que el archivo existe
        if not ruta.exists():
            raise FileNotFoundError(f"El archivo no existe: {ruta}")
            
        # Verificar que el archivo no esté vacío
        if ruta.stat().st_size == 0:
            raise ValueError(f"El archivo está vacío: {ruta}")
            
        # Seleccionar el motor apropiado según la extensión
        file_ext = ruta.suffix.lower()
        
        if file_ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            # Usar openpyxl para archivos .xlsx y formatos modernos
            print(f"  [INFO] Leyendo archivo {ruta.name} con motor openpyxl...")
            return pd.read_excel(ruta, sheet_name=hoja, engine='openpyxl')
        elif file_ext == '.xls':
            # Usar xlrd 2.0.1+ para archivos .xls modernos
            print(f"  [INFO] Leyendo archivo {ruta.name} con motor xlrd...")
            return pd.read_excel(ruta, sheet_name=hoja, engine='xlrd')
        else:
            raise ValueError(f"Formato de archivo no soportado: {file_ext}")
        
    except Exception as e:
        print(f"\n[ERROR] Error al leer el archivo {ruta.name}:")
        print(f"   Tipo de error: {type(e).__name__}")
        print(f"   Mensaje: {str(e)}")
        print(f"   Ruta completa: {ruta.absolute()}")
        print(f"   Tamaño: {ruta.stat().st_size/1024:.2f} KB" if ruta.exists() else "   El archivo no existe")
        raise


def detectar_mes_archivo(archivo: Path) -> str:
    """Detecta el mes del archivo FOCUS"""
    nombre = archivo.name.lower()
    
    print(f"  Analizando archivo: {archivo.name}")
    
    # Buscar por número de mes (01-12)
    for i, mes in enumerate(MESES, 1):
        if f"_{i:02d}_" in nombre or f"_{i:02d}." in nombre:
            print(f"  Mes detectado por número ({i:02d}): {mes}")
            return mes
    
    # Buscar por nombre de mes en el nombre del archivo
    for mes in MESES:
        if mes.lower() in nombre:
            print(f"  Mes detectado por nombre: {mes}")
            return mes
    
    # Si no se encuentra, intentar leer del contenido (todas las hojas)
    try:
        if str(archivo).lower().endswith('.xls'):
            # Para archivos .xls usar xlrd si está disponible
            if XLRD_AVAILABLE:
                wb = xlrd.open_workbook(str(archivo), on_demand=True)
                ws = wb.sheet_by_index(0)
                try:
                    valor_b5 = ws.cell(4, 1).value  # B5 es fila 4, columna 1 (0-based)
                except Exception:
                    valor_b5 = ""
            else:
                valor_b5 = ""
        else:
            # Para archivos .xlsx usar openpyxl y revisar todas las hojas
            wb = load_workbook(archivo, read_only=True, data_only=True)
            try:
                # Revisar B5 y H7 en todas las hojas hasta encontrar un mes válido
                for ws in wb.worksheets:
                    try:
                        valor_b5 = ws['B5'].value
                    except Exception:
                        valor_b5 = None
                    if valor_b5:
                        for mes in MESES:
                            if mes.lower() in str(valor_b5).lower():
                                print(f"  Mes detectado por B5 en hoja '{ws.title}': {mes}")
                                return mes
                    try:
                        valor_h7 = ws['H7'].value
                    except Exception:
                        valor_h7 = None
                    if valor_h7:
                        for mes in MESES:
                            if mes.lower() in str(valor_h7).lower():
                                print(f"  Mes detectado por H7 en hoja '{ws.title}': {mes}")
                                return mes
            finally:
                wb.close()
            
    except Exception as e:
        print(f"  [WARN] No se pudo leer el contenido del archivo para detectar el mes: {e}")
    
    # Default basado en el nombre del archivo
    if 'julio' in nombre:
        return "Julio"
    elif 'junio' in nombre:
        return "Junio"
    elif 'mayo' in nombre:
        return "Mayo"
    
    print(f"  Usando mes por defecto: Junio")
    return "Junio"  # Default


def obtener_trm_detallada(directorio: Path) -> dict:
    """
    Obtiene TRM detallada (usd, eur, fecha) desde trm_config.
    
    Args:
        directorio: Directorio base donde buscar la configuración
        
    Returns:
        dict: Diccionario con las claves: usd, eur, fecha, actualizado_por, origen
    """
    try:
        from trm_config import load_trm
        trm_data = load_trm()
        
        if not trm_data or not isinstance(trm_data, dict):
            raise ValueError("Datos de TRM no válidos")
            
        print(f"  TRM cargada desde trm_config: {trm_data.get('usd')} USD, {trm_data.get('eur')} EUR")
        return trm_data
        
    except Exception as e:
        print(f"  [ERROR] No se pudo cargar la TRM: {str(e)}")
        # Valores por defecto en caso de error
        return {
            'usd': 4000.0,
            'eur': 4500.0,
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'actualizado_por': 'sistema',
            'origen': 'error',
            'error': str(e)
        }

def obtener_trm(directorio: Path) -> Optional[float]:
    """Compatibilidad: devuelve sólo EUR si está disponible."""
    data = obtener_trm_detallada(directorio)
    return float(data['eur']) if data and 'eur' in data else None


def actualizar_celdas_trm(ws, base_dir: Path, mes_actual: Optional[str] = None) -> bool:
    """
    Actualiza todas las celdas relacionadas con TRM en la hoja de trabajo.
    
    Args:
        ws: La hoja de trabajo de Excel
        base_dir: Directorio base donde buscar la configuración de TRM
        mes_actual: Mes actual para mostrar en formato "Junio (Mes)"
        
    Returns:
        bool: True si la actualización fue exitosa, False en caso contrario
    """
    print("\nActualizando celdas con valores de TRM...")
    
    try:
        # Obtener la información TRM una sola vez
        trm_info = obtener_trm_detallada(base_dir)
        
        if not trm_info:
            print("  [ERROR] No se pudo obtener la información TRM")
            return False
            
        # Extraer valores TRM
        trm_eur = trm_info.get('eur')
        trm_usd = trm_info.get('usd')
        trm_fecha = trm_info.get('fecha')
        
        if not all([trm_eur, trm_usd, trm_fecha]):
            print(f"  [ERROR] Información TRM incompleta: EUR={trm_eur}, USD={trm_usd}, Fecha={trm_fecha}")
            return False
            
        print(f"  TRM actual: EUR={trm_eur}, USD={trm_usd}, Fecha={trm_fecha}")
        
        # 1. Actualizar tabla de tipo de cambio con nombres de mes
        # Área TIPO DE CAMBIO - CIERRE (generalmente alrededor de E5:G8)
        if mes_actual:
            try:
                # Buscar las celdas que contienen los meses
                # Formato: "Junio (Mes)" con valor 4.777,18
                mes_anterior_idx = MESES.index(mes_actual) - 1 if mes_actual in MESES else 0
                mes_anterior = MESES[mes_anterior_idx % 12]
                
                # Calcular promedio
                promedio_trm = trm_eur
                
                print(f"  [TRM] Mes actual: {mes_actual}, Mes anterior: {mes_anterior}")
                print(f"  [TRM] Promedio TRM EUR: {promedio_trm}")
                
            except Exception as e:
                print(f"  [WARN] Error calculando TRMs históricas: {e}")
        
        # 2. Actualizar tabla de conversión (generalmente en área derecha)
        # FECHA / DÓLAR TRM / EURO/PESO COL
        try:
            # Buscar y actualizar la tabla de conversión
            # Típicamente se encuentra alrededor de las columnas Q-S
            fecha_obj = datetime.strptime(trm_fecha, '%Y-%m-%d %H:%M:%S') if isinstance(trm_fecha, str) else datetime.now()
            fecha_formato = fecha_obj.strftime('%m/%d/%Y')
            
            print(f"  [TRM] Fecha formateada para tabla: {fecha_formato}")
            print(f"  [TRM] USD: {trm_usd}, EUR: {trm_eur}")
            
        except Exception as e:
            print(f"  [WARN] Error formateando fecha para tabla: {e}")
        
        # 3. Actualizar celdas de encabezado (Q3, Q4, P3, P4) 
        # Q3: Fecha actual del sistema
        # P3, P4: No modificar (dejar como estaban)
        # Q4: Valores TRM EUR
        try:
            # Q3: Fecha actual del sistema en formato dd/mm/yyyy
            fecha_actual = datetime.now()
            ws['Q3'].value = fecha_actual
            # Aplicar formato de fecha dd/mm/yyyy
            try:
                ws['Q3'].number_format = 'dd/mm/yyyy'
            except:
                pass
            print(f"  [OK] Actualizada Q3 con fecha actual: {fecha_actual.strftime('%d/%m/%Y')}")
            
            # P3: No modificar - dejar como está
            print(f"  [INFO] P3: Dejando valor original sin modificar")
            
            # P4: No modificar - dejar como está
            print(f"  [INFO] P4: Dejando valor original sin modificar")
            
            # Q4: Valor TRM EUR
            cell_q4 = ws['Q4']
            if cell_q4.data_type != 'f':  # Solo si no es fórmula
                cell_q4.value = trm_eur
                aplicar_formato_numero(ws, "Q4")
                print(f"  [OK] Actualizada Q4 con TRM EUR: {trm_eur}")
            else:
                print(f"  [LOCK] Q4: Se preserva la fórmula existente: {cell_q4.value}")
                
        except Exception as e:
            print(f"  [WARN] Error actualizando celdas de encabezado: {e}")
        
        # 4. Actualizar celdas de pie de página (Q42, R42, S42)
        try:
            # Actualizar fecha (Q42) con la fecha del archivo TRM
            ws['Q42'].value = trm_fecha
            print(f"  [OK] Actualizada Q42 con fecha: {trm_fecha}")
            
            # Actualizar USD (R42)
            ws['R42'].value = trm_usd
            aplicar_formato_numero(ws, "R42")
            print(f"  [OK] Actualizada R42 con TRM USD: {trm_usd}")
            
            # Actualizar EUR (S42)
            ws['S42'].value = trm_eur
            aplicar_formato_numero(ws, "S42")
            print(f"  [OK] Actualizada S42 con TRM EUR: {trm_eur}")
                
            # Asegurar que J7 tenga el valor correcto
            print(f"  [INFO] J7: preservando TRM cierre del modelo: {ws['J7'].value}")
                
            print("  [OK] Actualización de celdas TRM completada")
            return True
            
        except Exception as e:
            print(f"  [ERROR] Error actualizando celdas de pie de página: {e}")
            return False
            
    except Exception as e:
        print(f"  [ERROR] Error en el proceso de actualización de TRM: {e}")
        import traceback
        print(f"  Detalles del error: {traceback.format_exc()}")
        return False


def procesar_balance(df: pd.DataFrame) -> float:
    """
    1. BALANCE: Suma las cuentas especificadas
    Retorna el total en la unidad original (no dividido)
    """
    print("\n=== PROCESANDO BALANCE ===")

    # Intentar detectar columnas exactas usadas en tus archivos
    possible_account_cols = ["Número cuenta", "Numero cuenta", "Cuenta", "Número de cuenta", "Nro cuenta"]
    possible_saldo_cols = ["Saldo AAF variación", "Saldo AAF variacion", "Saldo", "Saldo periodo desviación", "Saldo periodo desviacion", "Libro mayor Saldo", "Saldo AAF"]

    col_cuenta = None
    col_saldo = None

    for c in possible_account_cols:
        if c in df.columns:
            col_cuenta = c
            break
    if col_cuenta is None:
        # fallback: primera columna no numérica
        for c in df.columns:
            # si la columna contiene strings (o tiene dtype object) asumimos que puede ser la cuenta
            is_object_dtype: bool = df[c].dtype == object
            has_strings: bool = df[c].apply(lambda x: isinstance(x, str)).any() == True
            if is_object_dtype or has_strings:
                col_cuenta = c
                break

    for c in possible_saldo_cols:
        if c in df.columns:
            col_saldo = c
            break
    if col_saldo is None:
        # fallback: columna con más valores numéricos
        numeric_counts: List[Tuple[Any, int]] = []
        for c in df.columns:
            def is_valid_number(x):
                if pd.isna(x):
                    return False
                return pd.api.types.is_number(x)
            is_numeric = df[c].apply(is_valid_number)
            count = int(is_numeric.sum())  # type: ignore
            numeric_counts.append((c, count))
        # Sort by count only, using negative count to get descending order
        # This avoids tuple comparison which causes type errors with Hashable column names
        numeric_counts = sorted(numeric_counts, key=lambda x: -x[1])
        if numeric_counts and numeric_counts[0][1] > 0:
            col_saldo = numeric_counts[0][0]
        else:
            # por último, tomar la última columna
            col_saldo = df.columns[-1]

    if col_cuenta is None or col_saldo is None:
        raise ValueError("No se pudieron detectar las columnas de cuenta o saldo en el archivo de balance. Columnas encontradas: " + str(list(df.columns)))

    print(f"  Columna cuenta: {col_cuenta}")
    print(f"  Columna saldo: {col_saldo}")

    # Normalizar columnas: eliminar espacios extra y convertir a str para búsqueda
    df = df.copy()
    df[col_cuenta] = df[col_cuenta].astype(str).str.strip().fillna("")
    # Reemplazar espacios no-break y caracteres invisibles
    df[col_cuenta] = df[col_cuenta].apply(lambda s: re.sub(r'\s+', ' ', s, flags=re.UNICODE).strip())

    total = 0.0
    coincidencias = []
    cuentas_procesadas = set()  # Para evitar duplicados

    # Primero intentar coincidencias exactas con la lista CUENTAS_BALANCE
    for cuenta in CUENTAS_BALANCE:
        # Solo coincidencia exacta, sin búsqueda parcial para evitar falsos positivos
        mask_exact = df[col_cuenta].str.strip().str.lower() == str(cuenta).strip().lower()
        if mask_exact.any():
            # Verificar que el valor sea numérico antes de sumar
            valores = df.loc[mask_exact, col_saldo].apply(to_float)
            # Filtrar valores extremadamente grandes que podrían ser errores
            valores = valores[abs(valores) < 1e12]  # Filtrar valores mayores a 1 billón
            subtotal = valores.sum()
            
            if abs(subtotal) > 0:  # Solo sumar si el valor es significativo
                # Verificar que la cuenta no haya sido procesada antes
                cuenta_key = str(cuenta).strip().lower()
                if cuenta_key not in cuentas_procesadas:
                    total += subtotal
                    cuentas_procesadas.add(cuenta_key)
                    coincidencias.append((cuenta, subtotal, "exacta"))
                    print(f"  [EXACTA] {cuenta}: {subtotal:,.2f}")
                else:
                    print(f"  [DUPLICADA] {cuenta}: {subtotal:,.2f} (ya procesada)")

    # Buscar específicamente por los totales que necesitamos
    print("  Buscando totales específicos en el nombre de cuenta...")
    
    # Definir patrones de cuentas de total que queremos incluir
    patrones_totales = [
        'TOTAL ACTIVO', 'TOTAL DEL ACTIVO', 'TOTAL ACTIVOS',
        'TOTAL PASIVO', 'TOTAL DEL PASIVO', 'TOTAL PASIVOS',
        'TOTAL PATRIMONIO', 'TOTAL DEL PATRIMONIO'
    ]
    
    # Inicializar subtotal para totales
    subtotal_total = 0.0
    
    # Primero, verificar si hay una fila que sea exactamente "TOTAL"
    mask_total = (df[col_cuenta].str.strip().str.upper() == 'TOTAL')
    if mask_total.any():
        subtotal = df.loc[mask_total, col_saldo].apply(to_float).sum()
        if abs(subtotal) > 0 and abs(subtotal) < 1e12:  # Filtro para valores extremos
            subtotal_total += subtotal
            coincidencias.append(("TOTAL", subtotal, "total_general"))
            print(f"  [TOTAL GENERAL] {subtotal:,.2f}")
    
    # Si no encontramos un TOTAL general, buscar por patrones específicos
    if subtotal_total == 0:
        for patron in patrones_totales:
            mask = df[col_cuenta].str.strip().str.upper() == patron.upper()
            if mask.any():
                subtotal = df.loc[mask, col_saldo].apply(to_float).sum()
                if abs(subtotal) > 0 and abs(subtotal) < 1e12:  # Filtro para valores extremos
                    # Verificar que el patrón no haya sido procesado antes
                    patron_key = str(patron).strip().lower()
                    if patron_key not in cuentas_procesadas:
                        subtotal_total += subtotal
                        cuentas_procesadas.add(patron_key)
                        coincidencias.append((patron, subtotal, "total_especifico"))
                        print(f"  [TOTAL] {patron}: {subtotal:,.2f}")
    
    # Solo sumar si encontramos algún total específico
    if abs(subtotal_total) > 0 and abs(subtotal_total) < 1e12:  # Validación adicional
        total += subtotal_total
    else:
        print("  [WARN] Se ignoró un total con valor sospechoso:", subtotal_total)
    
    # Buscar también filas que NO contengan "TOTAL" pero que sean cuentas válidas
    print("\n  === DETALLE DE CÁLCULO DEL BALANCE ===")
    print("  ---------------------------------")
    print("  CUENTAS ENCONTRADAS Y SUS VALORES:")
    
    # Mostrar coincidencias exactas
    if any(t[2] == "exacta" for t in coincidencias):
        print("\n  [COINCIDENCIAS EXACTAS]")
        for cuenta, valor, tipo in coincidencias:
            if tipo == "exacta":
                print(f"  - {cuenta}: {valor:,.2f}")
    
    # Mostrar totales encontrados
    if any(t[2] in ["total_general", "total_especifico"] for t in coincidencias):
        print("\n  [TOTALES ENCONTRADOS]")
        for cuenta, valor, tipo in coincidencias:
            if tipo in ["total_general", "total_especifico"]:
                print(f"  - {cuenta}: {valor:,.2f}")
    
    # Buscar filas sin 'TOTAL' pero con cuentas válidas que comiencen con "0080"
    print("\n  [BUSCANDO CUENTAS 0080]")
    # Buscar cuentas que comienzan con "0080" (sin incluir las que contienen "TOTAL")
    mask_no_total = ~df[col_cuenta].str.contains('total', case=False, na=False)
    mask_cuentas_0080 = df[col_cuenta].str.contains(r'^0080\.', case=False, na=False, regex=True)
    mask_final = mask_no_total & mask_cuentas_0080
    
    if mask_final.any():
        cuentas_encontradas = []
        for idx, row in df[mask_final].iterrows():
            cuenta_str = str(row[col_cuenta]).strip()
            valor = to_float(row[col_saldo])
            if valor != 0:
                # Verificar que la cuenta no haya sido procesada antes
                cuenta_key = cuenta_str.lower()
                if cuenta_key not in cuentas_procesadas:
                    cuentas_encontradas.append((cuenta_str, valor))
                    total += valor
                    cuentas_procesadas.add(cuenta_key)
                    coincidencias.append((cuenta_str, valor, "no_total_0080"))
                else:
                    print(f"  [DUPLICADA] {cuenta_str}: {valor:,.2f} (ya procesada)")
        
        if cuentas_encontradas:
            print("  Se encontraron las siguientes cuentas 0080:")
            for cuenta, valor in cuentas_encontradas:
                print(f"  - {cuenta}: {valor:,.2f}")
    
    # Búsqueda por patrones numéricos (solo si no hubo coincidencias previas)
    if not coincidencias:
        print("\n  [BÚSQUEDA POR PATRONES NUMÉRICOS]")
        print("  No se encontraron coincidencias con CUENTAS_BALANCE. Buscando patrones alternativos...")
        patrones = ["0080.43002", "0080.01010", "43001", "43008", "43042"]
        cuentas_patron = []
        
        for idx, row in df.iterrows():
            cuenta_val = str(row[col_cuenta])
            if any(p in cuenta_val for p in patrones):
                valor = to_float(row[col_saldo])
                if valor != 0:
                    # Verificar que la cuenta no haya sido procesada antes
                    cuenta_key = cuenta_val.lower()
                    if cuenta_key not in cuentas_procesadas:
                        cuentas_patron.append((cuenta_val, valor))
                        total += valor
                        cuentas_procesadas.add(cuenta_key)
                        coincidencias.append((cuenta_val, valor, "patron"))
                    else:
                        print(f"  [DUPLICADA] {cuenta_val}: {valor:,.2f} (ya procesada)")
        
        if cuentas_patron:
            print("  Se encontraron las siguientes cuentas por patrón:")
            for cuenta, valor in cuentas_patron:
                print(f"  - {cuenta}: {valor:,.2f}")
    
    # Mostrar resumen final
    print("\n  === RESUMEN DEL CÁLCULO ===")
    print("  ------------------------")
    if coincidencias:
        print(f"  Total de cuentas encontradas: {len(coincidencias)}")
        print(f"  Suma total calculada: {total:,.2f}")
    else:
        print("  [ADVERTENCIA] No se encontraron cuentas para calcular el balance")
    print("  ============================\n")
    print(f"TOTAL BALANCE: {total:,.2f}")
    return total


def procesar_situacion(df: pd.DataFrame) -> Tuple[float, float, float, float]:
    """
    Extrae de la situación de cuentas los cobros de S01 - COBROS DE CLIENTES,
    separados en Vencida / No Vencida (lectura directa, sin proporción).

    Busca la fila donde COD.CATEGORÍA 1 == 'S01 - COBROS DE CLIENTES' y la
    SUBCUENTA/CUENTA OBJETO == '01010', y de esa fila lee dos columnas
    independientes: el saldo vencido del mes y el saldo no vencido del mes.

    Si no se encuentran columnas separadas de vencida/no vencida, cae en un
    fallback de último recurso devolviendo (total, 0.0, total, acumulado) -
    señal para que el código principal use el split proporcional D14/H14
    SOLO como último recurso.

    Retorna: (cobros_vencida_mes, cobros_no_vencida_mes, cobros_total_mes, cobros_acumulado)
    Los cuatro valores se devuelven SIEMPRE con signo negativo (los cobros
    son un movimiento que resta de la cartera, nunca una entrada positiva).
    """
    print("\n=== PROCESANDO SITUACIÓN ===")

    # ── Detectar columnas clave ─────────────────────────────────────────────
    col_saldo_mes         = None
    col_saldo_acum        = None
    col_cat1              = None   # COD.CATEGORÍA 1  -> identifica S01
    col_subcuenta         = None   # SUBCUENTA / CUENTA OBJETO -> identifica 01010
    col_saldo_vencido     = None   # Saldo vencido del mes (split real)
    col_saldo_no_vencido  = None   # Saldo no vencido del mes (split real)

    for col in df.columns:
        col_norm = normalize_text(str(col))

        if (col_saldo_mes is None and 'saldo' in col_norm and 'mes' in col_norm
                and 'acum' not in col_norm and 'vencid' not in col_norm):
            col_saldo_mes = col
        if (col_saldo_acum is None and 'saldo' in col_norm
                and ('acum' in col_norm or 'acumulado' in col_norm) and 'vencid' not in col_norm):
            col_saldo_acum = col
        if col_cat1 is None and 'categor' in col_norm and '1' in col_norm:
            col_cat1 = col
        if col_subcuenta is None and ('subcuenta' in col_norm or 'cuenta objeto' in col_norm or 'objeto' in col_norm):
            col_subcuenta = col

    # ── Detectar columnas de split Vencida / No Vencida (Cambio 3) ───────────
    # Pasada 1: exigir que el nombre incluya "mes" para no chocar con acumuladas
    for col in df.columns:
        col_norm      = normalize_text(str(col))
        es_no_vencido = 'no vencid' in col_norm
        es_vencido    = 'vencid' in col_norm and not es_no_vencido
        es_acum       = 'acum' in col_norm or 'acumulado' in col_norm

        if es_acum:
            continue
        if es_vencido and 'mes' in col_norm and col_saldo_vencido is None:
            col_saldo_vencido = col
        if es_no_vencido and 'mes' in col_norm and col_saldo_no_vencido is None:
            col_saldo_no_vencido = col

    # Pasada 2: sin exigir "mes", si no se encontraron en la pasada 1
    if col_saldo_vencido is None or col_saldo_no_vencido is None:
        for col in df.columns:
            col_norm      = normalize_text(str(col))
            es_no_vencido = 'no vencid' in col_norm
            es_vencido    = 'vencid' in col_norm and not es_no_vencido
            es_acum       = 'acum' in col_norm or 'acumulado' in col_norm

            if es_acum:
                continue
            if es_vencido and col_saldo_vencido is None:
                col_saldo_vencido = col
            if es_no_vencido and col_saldo_no_vencido is None:
                col_saldo_no_vencido = col

    # Fallbacks por posición (estructura del archivo Colombia Situación)
    if col_saldo_mes is None and len(df.columns) >= 9:
        col_saldo_mes = df.columns[8]
    if col_saldo_acum is None and len(df.columns) >= 12:
        col_saldo_acum = df.columns[11]
    if col_cat1 is None and len(df.columns) >= 5:
        col_cat1 = df.columns[4]   # COD.CATEGORÍA 1 es la col 4 (0-based) en el archivo real
    if col_subcuenta is None and len(df.columns) >= 6:
        col_subcuenta = df.columns[5]  # CUENTA OBJETO es col 5

    print(f"  Columna SALDO MES        : {col_saldo_mes}")
    print(f"  Columna SALDO ACUMULADO  : {col_saldo_acum}")
    print(f"  Columna COD.CATEGORÍA 1  : {col_cat1}")
    print(f"  Columna SUBCUENTA/OBJ    : {col_subcuenta}")
    print(f"  Columna SALDO VENCIDO    : {col_saldo_vencido}")
    print(f"  Columna SALDO NO VENCIDO : {col_saldo_no_vencido}")

    cobros_total          = 0.0
    cobros_acum           = 0.0
    cobros_vencida_mes    = 0.0
    cobros_no_vencida_mes = 0.0
    encontrado            = False
    fila_encontrada        = None

    S01_NORM       = normalize_text('S01 - COBROS DE CLIENTES')
    TOTAL_S01_NORM = normalize_text('Total S01 - COBROS DE CLIENTES')

    # ── Estrategia 1 (FIX): buscar la fila de TOTAL real por texto exacto ────
    if col_cat1:
        for idx, row in df.iterrows():
            cat1_val = normalize_text(str(row.get(col_cat1, '') or '')).strip()
            if cat1_val == TOTAL_S01_NORM:
                fila_encontrada = row
                encontrado = True
                print(f"  [S01-TOTAL] Fila encontrada en idx={idx}")
                break

    # ── Estrategia 2 (fallback): "Total S01..." en cualquier celda ───────────
    if not encontrado:
        print("  [WARN] Estrategia 1 no encontró la fila. Intentando fallback por texto...")
        for idx, row in df.iterrows():
            if any(TOTAL_S01_NORM in normalize_text(str(v)) for v in row.values):
                fila_encontrada = row
                encontrado = True
                print(f"  [S01-fallback] Fila encontrada en idx={idx}")
                break

    # ── Estrategia 3 (último recurso): Total 01010 genérico ──────────────────
    if not encontrado:
        print("  [WARN] Fallback S01 tampoco encontró. Usando Total 01010 genérico...")
        TARGET_GENERICO = normalize_text('TOTAL 01010')
        for idx, row in df.iterrows():
            if any(TARGET_GENERICO in normalize_text(str(v)) for v in row.values):
                fila_encontrada = row
                encontrado = True
                print(f"  [S01-generico] Fila encontrada en idx={idx}")
                break

    if not encontrado or fila_encontrada is None:
        print("  [ERROR] No se encontró fila de cobros S01 en ninguna estrategia")
        return 0.0, 0.0, 0.0, 0.0

    row = fila_encontrada
    cobros_total = to_float(row[col_saldo_mes])  if col_saldo_mes  else 0.0
    cobros_acum  = to_float(row[col_saldo_acum]) if col_saldo_acum else 0.0

    if col_saldo_vencido is not None and col_saldo_no_vencido is not None:
        cobros_vencida_mes    = to_float(row[col_saldo_vencido])
        cobros_no_vencida_mes = to_float(row[col_saldo_no_vencido])
        print(f"  [SPLIT-REAL] Vencida={cobros_vencida_mes:,.0f}  No Vencida={cobros_no_vencida_mes:,.0f}")

        # Si no hay columna de total separada, derivarlo del split real
        if cobros_total == 0.0:
            cobros_total = cobros_vencida_mes + cobros_no_vencida_mes
    else:
        print("  [WARN] No se encontraron columnas separadas de Vencida/No Vencida.")
        print("  [WARN] Se devuelve solo el total; el split proporcional D14/H14 se "
              "usará como ÚLTIMO RECURSO en el código principal.")
        cobros_vencida_mes    = cobros_total
        cobros_no_vencida_mes = 0.0

    # Convertir a miles si viene en pesos (valor > 1e8 es señal clara de pesos)
    if abs(cobros_total) > 1e8:
        print(f"  [SCALE] Cobros detectados en pesos -> dividiendo /1000")
        cobros_total          = cobros_total          / 1000.0
        cobros_acum           = cobros_acum           / 1000.0
        cobros_vencida_mes    = cobros_vencida_mes    / 1000.0
        cobros_no_vencida_mes = cobros_no_vencida_mes / 1000.0

    # ── Normalización de signo (FIX) ──────────────────────────────────────────
    # Los cobros, sin importar si vinieron del split real o son solo un total
    # (camino que alimenta el fallback proporcional en el código principal),
    # SIEMPRE deben quedar como un movimiento NEGATIVO: representan dinero que
    # sale de la cartera y reduce el saldo, nunca una entrada positiva.
    # El archivo de Situación los entrega como montos cobrados (positivos),
    # así que aquí se invierte el signo antes de devolverlos. Esto garantiza
    # que tanto el split real como el fallback proporcional queden bien,
    # sin tener que tocar nada en procesar_y_actualizar_focus().
    cobros_vencida_mes    = -abs(cobros_vencida_mes)
    cobros_no_vencida_mes = -abs(cobros_no_vencida_mes)
    cobros_total          = -abs(cobros_total)
    cobros_acum           = -abs(cobros_acum)

    print(f"  [SIGN-FIX] Vencida={cobros_vencida_mes:,.3f}  No Vencida={cobros_no_vencida_mes:,.3f}")
    print(f"  COBROS VENCIDA MES (miles)    : {cobros_vencida_mes:,.3f}")
    print(f"  COBROS NO VENCIDA MES (miles) : {cobros_no_vencida_mes:,.3f}")
    print(f"  COBROS TOTAL MES (miles)      : {cobros_total:,.3f}")
    print(f"  COBROS ACUMULADO (miles)      : {cobros_acum:,.3f}")

    return cobros_vencida_mes, cobros_no_vencida_mes, cobros_total, cobros_acum

def _encontrar_fila_total_modelo(ws) -> Optional[int]:
    """
    Busca la ÚLTIMA fila del Modelo Deuda que:
    - Contenga la palabra 'Total' exacto en CUALQUIER columna (A hasta T)
    - Tenga un valor numérico grande en columna G (> 1e6)
    Se sobreescribe en cada match → al terminar queda la última = gran total.
    """
    from openpyxl.utils import column_index_from_string
    fila_total = None

    for row_idx in range(1, ws.max_row + 1):
        col_g = to_float(ws.cell(row=row_idx, column=column_index_from_string('G')).value)
        if col_g <= 1e6:
            continue  # no es fila de totales grandes, skip rápido

        # Revisar TODAS las celdas de la fila buscando 'total' exacto
        for col_idx in range(1, 20):  # columnas A hasta T
            val = str(ws.cell(row=row_idx, column=col_idx).value or '').strip().lower()
            if val == 'total':
                fila_total = row_idx  # se sobreescribe → siempre queda la última
                print(f"  [TOTAL candidato] fila {row_idx} col {col_idx}: G={col_g:,.0f}")
                break  # encontrado en esta fila, pasar a la siguiente

    if fila_total:
        print(f"  [TOTAL GENERAL] Última fila seleccionada: {fila_total}")
    else:
        print("  [ERROR] No se encontró fila 'Total' con valores grandes")

    return fila_total


def procesar_modelo_como_espana(ruta: Path) -> Dict[str, float]:
    print("\n=== PROCESANDO MODELO DEUDA COMO FUENTE DE ESPAÑA ===")
    resultado = {
        'h22': 0.0, 'd22': 0.0, 'f22': 0.0, 'usd_total': 0.0,
        'saldo_total_gran': 0.0, 'venc_30_gran': 0.0, 'suma_vencidos_gran': 0.0,
        'incobrable_gran': 0.0,
    }

    try:
        wb = load_workbook(ruta, data_only=True)
        ws_nombre = next(
            (h for h in wb.sheetnames if 'MODELO DEUDA' in h.strip().upper()),
            wb.sheetnames[0]
        )
        ws = wb[ws_nombre]
        print(f"  Hoja seleccionada: {ws_nombre}")

        from openpyxl.utils import column_index_from_string
        def gv(row, col_letter):
            return to_float(ws.cell(row=row, column=column_index_from_string(col_letter)).value)

        # ── Usar helper compartido para fila_total ───────────────────────────
        fila_total = _encontrar_fila_total_modelo(ws)

        # ── Recorrer filas para pesos, USD puro y TRM ───────────────────────
        fila_pesos    = None
        fila_usd_puro = None
        fila_trm      = None

        for row_idx in range(1, ws.max_row + 1):
            col_e   = str(ws.cell(row=row_idx, column=5).value or '').strip()
            col_f   = str(ws.cell(row=row_idx, column=6).value or '').strip()
            col_g   = to_float(ws.cell(row=row_idx, column=7).value)

            # Fila subtotal PESOS COL
            if 'moneda local' in col_e.lower() and col_g > 0:
                fila_pesos = row_idx
                print(f"  [PESOS] Encontrado en fila {row_idx}: G={col_g:,.0f}")

            # Fila subtotal USD puro
            if 'total moneda extranjera usd' in col_f.lower():
                fila_usd_puro = row_idx
                print(f"  [USD PURO] Encontrado en fila {row_idx}")

            # Fila TRM (conversión USD→COP)
            trm_val = to_float(col_f)
            if ('dólar' in col_e.lower() or 'dolar' in col_e.lower()) and trm_val > 1000:
                fila_trm = row_idx
                print(f"  [TRM] Encontrado en fila {row_idx}: TRM={trm_val}")

        if not all([fila_pesos, fila_usd_puro, fila_total]):
            print(f"  [ERROR] Filas clave faltantes: "
                  f"pesos={fila_pesos}, usd={fila_usd_puro}, total={fila_total}")
            return resultado

        # ── PESOS COL: H22, D22, F22 ────────────────────────────────────────
        G_pesos = gv(fila_pesos, 'G')
        H_pesos = gv(fila_pesos, 'H')
        I_pesos = gv(fila_pesos, 'I')
        J_pesos = gv(fila_pesos, 'J')
        K_pesos = gv(fila_pesos, 'K')
        L_pesos = gv(fila_pesos, 'L')
        M_pesos = gv(fila_pesos, 'M')
        N_pesos = gv(fila_pesos, 'N')
        vencido_pesos = I_pesos + J_pesos + K_pesos + L_pesos + M_pesos + N_pesos

        h22 = G_pesos       / 1000.0
        d22 = vencido_pesos / 1000.0
        f22 = H_pesos       / 1000.0
        print(f"  [PESOS] H22={h22:,.3f} | D22={d22:,.3f} | F22={f22:,.3f}")

        # ── USD puro ─────────────────────────────────────────────────────────
        usd_total = gv(fila_usd_puro, 'G')
        print(f"  [USD] usd_total={usd_total:,.3f} USD puros")

        # ── GRAN TOTAL (última fila 'Total') ─────────────────────────────────
        saldo_total_gran   = gv(fila_total, 'G')
        venc_30_gran       = gv(fila_total, 'I')
        venc_60_gran       = gv(fila_total, 'J')
        venc_90_gran       = gv(fila_total, 'K')
        venc_180_gran      = gv(fila_total, 'L')
        venc_360_gran      = gv(fila_total, 'M')
        venc_360p_gran     = gv(fila_total, 'N')
        incobrable_gran    = gv(fila_total, 'O')
        suma_vencidos_gran = (venc_60_gran + venc_90_gran + venc_180_gran
                              + venc_360_gran + venc_360p_gran)

        print(f"  [TOTAL G] Saldo={saldo_total_gran:,.0f} | Venc30={venc_30_gran:,.0f} | "
              f"SumaJ:N={suma_vencidos_gran:,.0f} | Incobrable={incobrable_gran:,.0f}")

        wb.close()

    except Exception as e:
        print(f"  [ERROR] No se pudo leer el archivo Modelo Deuda: {e}")
        import traceback; print(traceback.format_exc())
        return resultado

    resultado.update({
        'h22':                h22,
        'd22':                d22,
        'f22':                f22,
        'usd_total':          usd_total,
        'saldo_total_gran':   saldo_total_gran,
        'venc_30_gran':       venc_30_gran,
        'suma_vencidos_gran': suma_vencidos_gran,
        'incobrable_gran':    incobrable_gran,
        'trm_usd_archivo':    gv(fila_trm, 'F') if fila_trm else 0.0,
    })

    return resultado


def procesar_modelo_vencimiento(ruta: Path) -> float:
    """Lee deuda incobrable total de col O de la ÚLTIMA fila con 'Total' y G > 1e6."""
    print("\n=== PROCESANDO MODELO VENCIMIENTO (incobrable total) ===")
    try:
        wb = load_workbook(ruta, data_only=True)
        ws_nombre = next(
            (h for h in wb.sheetnames if 'MODELO DEUDA' in h.strip().upper()),
            wb.sheetnames[0]
        )
        ws = wb[ws_nombre]

        fila_total = _encontrar_fila_total_modelo(ws)

        if not fila_total:
            print("  [ERROR] No se encontró fila Total en Modelo Vencimiento")
            wb.close()
            return 0.0

        from openpyxl.utils import column_index_from_string
        incobrable = to_float(
            ws.cell(row=fila_total, column=column_index_from_string('O')).value
        )
        print(f"  [OK] Incobrable fila {fila_total} col O = {incobrable:,.2f}")
        wb.close()
        return incobrable

    except Exception as e:
        print(f"  [ERROR] {e}")
        return 0.0
    
def leer_celda_fija_situacion(ruta_situacion: Path) -> float:
    """
    Lee el TOTAL COBROS S01 buscando dinámicamente tanto la FILA
    ('Total S01 - COBROS DE CLIENTES') como la COLUMNA ('SALDO MES'),
    porque ambas cambian de posición según el mes (algunos archivos
    no traen columna DIVISIÓN y/o agregan columna 'SALDO MES -1').
    """
    print("\n=== LEYENDO TOTAL S01 DE SITUACIÓN (columna dinámica) ===")
    try:
        wb_sit = load_workbook(ruta_situacion, data_only=True)
        ws_sit = wb_sit.worksheets[0]

        # 1. Encontrar la columna real de "SALDO MES" (NO "SALDO MES -1")
        col_saldo_mes = None
        for row in ws_sit.iter_rows(min_row=1, max_row=10):
            for cell in row:
                val_raw = str(cell.value or '')
                val_norm = normalize_text(val_raw)
                if 'saldo mes' in val_norm and '-1' not in val_raw:
                    col_saldo_mes = cell.column
                    break
            if col_saldo_mes:
                break

        if not col_saldo_mes:
            print("  [WARN] No se encontró encabezado 'SALDO MES', usando columna I (9) por defecto")
            col_saldo_mes = 9
        else:
            print(f"  [OK] Columna SALDO MES detectada: {get_column_letter(col_saldo_mes)}")

        # 2. Encontrar la fila "Total S01 - COBROS DE CLIENTES"
        KEYWORDS_TOTAL = ['total s01 - cobros de clientes', 'total s01']

        for row_idx in range(1, ws_sit.max_row + 1):
            for col_idx in range(1, 15):  # rango ampliado por seguridad
                val = normalize_text(
                    str(ws_sit.cell(row=row_idx, column=col_idx).value or '')
                )
                if any(kw in val for kw in KEYWORDS_TOTAL):
                    resultado = to_float(
                        ws_sit.cell(row=row_idx, column=col_saldo_mes).value
                    )
                    col_letra = get_column_letter(col_saldo_mes)
                    print(f"  [OK] Fila {row_idx}: '{ws_sit.cell(row=row_idx, column=col_idx).value}'"
                          f" → {col_letra}{row_idx} = {resultado:,.3f}")
                    wb_sit.close()
                    return resultado

        print("  [WARN] No se encontró 'Total S01' en ninguna fila")
        wb_sit.close()
        return 0.0

    except Exception as e:
        print(f"  [ERROR] No se pudo leer Total S01 de Situación: {e}")
        return 0.0
    
def aplicar_formato_numero(ws, celda: str):
    """Aplica formato de número y alineación correcta a una celda"""
    try:
        cell = ws[celda]
        if hasattr(cell, 'number_format'):
            cell.number_format = '#,##0.00'
        if hasattr(cell, 'alignment'):
            from openpyxl.styles import Alignment
            cell.alignment = Alignment(horizontal='right', vertical='center')
    except Exception as e:
        print(f"  [WARN] Error aplicando formato a {celda}: {str(e)}")

def limpiar_celda_antes_de_pegar(ws, celda: str, descripcion: str = ""):
    """Limpia una celda antes de pegar un nuevo valor (BORRAR EL RESULTADO ANTES DE PEGAR)"""
    try:
        cell = ws[celda]
        if cell.data_type != 'f':  # Solo limpiar si NO es una fórmula
            valor_anterior = cell.value
            cell.value = None
            print(f"  [CLEAN] {celda} limpiada antes de pegar {descripcion} (valor anterior: {valor_anterior})")
            return True
        else:
            print(f"  [LOCK] {celda}: PRESERVANDO fórmula existente: {cell.value}")
            return False
    except Exception as e:
        print(f"  [WARN] Error limpiando celda {celda}: {str(e)}")
        return False

def validar_datos(ws, ws_values, total_balance: float, total_vencido: float) -> None:
    """Valida totales clave contra especificaciones del PDF."""
    try:
        h22 = to_float(ws_values["H22"].value) if ws_values else to_float(ws["H22"].value)
    except Exception:
        h22 = to_float(ws["H22"].value)
    try:
        q15 = to_float(ws_values["Q15"].value) if ws_values else to_float(ws["Q15"].value)
    except Exception:
        q15 = to_float(ws["Q15"].value)

    if abs(h22 - q15) > 0.01:
        print(f"[ERROR] ERROR: H22 ({h22:,.2f}) != Q15 ({q15:,.2f})")

    try:
        d22 = to_float(ws_values["D22"].value) if ws_values else to_float(ws["D22"].value)
    except Exception:
        d22 = to_float(ws["D22"].value)
    if abs(d22 - (total_vencido / 1000.0)) > 0.01:
        print("[ERROR] ERROR: D22 no coincide con total vencido del modelo")

    try:
        f22 = to_float(ws_values["F22"].value) if ws_values else to_float(ws["F22"].value)
    except Exception:
        f22 = to_float(ws["F22"].value)
    total_no_vencido = (total_balance - total_vencido) / 1000.0
    if abs(f22 - total_no_vencido) > 0.01:
        print("[ERROR] ERROR: F22 no coincide con total no vencido del modelo")

def escribir_celda(ws, celda: str, valor, como_numero=True, preservar_formula=True, forzar_sobrescribir=False):
    """Escribe un valor en una celda con formato apropiado. NUNCA sobrescribe fórmulas existentes."""
    try:
        # Solo usar openpyxl (el código ya no usa xlwt para escribir)
        cell = ws[celda]
        # NUNCA sobrescribir si la celda tiene una fórmula
        if cell.data_type == 'f':
            print(f"  [LOCK] PRESERVANDO fórmula en {celda}: {cell.value}")
            return False
        # Solo escribir si NO es una fórmula
        if valor is None:
            cell.value = None
        elif como_numero:
            try:
                cell.value = float(valor)
                # Aplicar formato y alineación correcta
                aplicar_formato_numero(ws, celda)
            except (ValueError, TypeError):
                cell.value = valor
        else:
            cell.value = valor
        return True
    except Exception as e:
        print(f"  [WARN] Error al escribir en celda {celda}: {str(e)}")
        return False


def insertar_dato_entrada(ws, celda: str, valor, descripcion: str = "", forzar_sobrescribir: bool = False):
    """
    Inserta un dato de entrada (valor fijo) en celdas específicas.
    Estas son las celdas de datos de entrada que deben tener valores fijos.
    
    Args:
        ws: Worksheet
        celda: Referencia de celda (ej: "A1")
        valor: Valor a escribir
        descripcion: Descripción del dato para el log
        forzar_sobrescribir: Si True, sobrescribe incluso fórmulas existentes
    """
    try:
        cell = ws[celda]
        
        # Verificar si la celda ya tiene una fórmula
        if cell.data_type == 'f' and cell.value and not forzar_sobrescribir:
            print(f"  [LOCK] PRESERVANDO fórmula existente en {celda}: {cell.value}")
            print(f"  [DOC] {descripcion} - No se sobrescribe fórmula existente")
            return True
        
        # Verificar si la celda contiene una referencia al PDF que debería mantenerse
        if cell.value and not forzar_sobrescribir:
            valor_actual = str(cell.value)
            # Si la celda contiene una referencia al PDF, agregar el valor calculado como comentario
            if 'FORMATO DEUDA' in valor_actual.upper() or '@' in valor_actual:
                print(f"  [DOC] {descripcion} - Celda {celda} contiene referencia a PDF: {valor_actual}")
                # En lugar de sobrescribir, agregamos el valor calculado como parte del texto
                nuevo_valor = f"{valor_actual} (Valor calculado: {valor})"
                cell.value = nuevo_valor
                print(f"  [DOC] {descripcion} - Actualizado en {celda}: {nuevo_valor}")
                return True
        
        print(f"  [DOC] {descripcion} - Insertando valor fijo en {celda}: {valor}")
        
        if valor is None:
            cell.value = None
        else:
            try:
                # Intentar convertir a número solo si es posible
                if isinstance(valor, (int, float)) or (isinstance(valor, str) and valor.replace('.', '').replace('-', '').isdigit()):
                    cell.value = float(valor)
                    # Aplicar formato y alineación correcta
                    aplicar_formato_numero(ws, celda)
                else:
                    # Para texto (como meses), no aplicar formato numérico
                    cell.value = valor
            except (ValueError, TypeError):
                cell.value = valor
        
        print(f"  [DOC] {descripcion} - Celda {celda}: {valor}")
        return True
        
    except Exception as e:
        print(f"  [WARN] Error insertando {descripcion} en {celda}: {str(e)}")
        return False


def insertar_formula(ws, celda: str, formula: str, descripcion: str = ""):
    """
    Inserta una fórmula en una celda específica, pero solo si no tiene una fórmula existente.
    
    Args:
        ws: Worksheet
        celda: Referencia de celda (ej: "A1")
        formula: Fórmula a insertar (ej: "=Q16/1000")
        descripcion: Descripción de la fórmula para el log
    """
    try:
        cell = ws[celda]
        
        # Verificar si la celda ya tiene una fórmula
        if cell.data_type == 'f' and cell.value:
            print(f"  [LOCK] PRESERVANDO fórmula existente en {celda}: {cell.value}")
            print(f"  [FORMULA] {descripcion} - No se sobrescribe fórmula existente")
            return True
        
        print(f"  [FORMULA] {descripcion} - Insertando fórmula en {celda}: {formula}")
        
        # Insertar la fórmula solo si no hay una existente
        cell.value = formula
        
        print(f"  [FORMULA] {descripcion} - Fórmula insertada en {celda}")
        return True
        
    except Exception as e:
        print(f"  [WARN] Error insertando fórmula {descripcion} en {celda}: {str(e)}")
        return False
       
def rotar_finales_a_iniciales(ws_modelo: Worksheet, ruta_archivo: Path) -> None:
    """
    Lee los valores FINALES del archivo guardado (data_only=True) y los
    escribe como INICIALES en ws_modelo (que está abierto para edición).

    NOTA: el bloque ACUM (filas 35 y 44, "Inicial ACUM") ya NO se alimenta de
    su "Final ACUM" correspondiente (filas 43 y 50). En la plantilla actual
    esas celdas son fórmulas que apuntan directo a su par del bloque MES
    (fila 14 y 23 respectivamente) — "inicial a inicial", no "final a
    inicial". Por eso se actualizan solas en cuanto la rotación de MES deja
    D14/F14/H14 y D23/F23/H23 correctos; no hace falta (ni se debe) tocarlas
    aquí.
    """
    print("\n=== ROTANDO FINALES A INICIALES ===")

    wb_readonly = load_workbook(ruta_archivo, data_only=True)
    ws_readonly = None
    for sn in wb_readonly.sheetnames:
        if "MODELO" in sn.strip().upper():
            ws_readonly = wb_readonly[sn]
            break
    if ws_readonly is None:
        ws_readonly = wb_readonly.active

    columnas = ['D', 'F', 'H', 'J', 'L', 'N']

    pares_rotacion = [
        (22, 14),   # Deuda bruta Final MES -> Inicial MES
        (29, 23),   # Provision acumulada Final MES -> Dotaciones Inicial MES
        (43, 35),   # Cobros/Fact/Vencidos ACUM Final -> Inicial ACUM
        (50, 44),   # Dotaciones ACUM Final -> Dotaciones ACUM Inicial
    ]

    for fila_origen, fila_destino in pares_rotacion:
        for col in columnas:
            celda_origen  = f"{col}{fila_origen}"
            celda_destino = f"{col}{fila_destino}"
            try:
                valor = ws_readonly[celda_origen].value
                if valor is None:
                    continue

                cell_destino = ws_modelo[celda_destino]
                if cell_destino.data_type == 'f':
                    print(f"  [LOCK] {celda_destino}: formula preservada")
                    continue

                cell_destino.value = float(valor)
                aplicar_formato_numero(ws_modelo, celda_destino)
                print(f"  [ROT] {celda_origen} ({float(valor):,.2f}) -> {celda_destino}")

            except (ValueError, TypeError) as e:
                print(f"  [WARN] No se pudo rotar {celda_origen} -> {celda_destino}: {e}")
            except Exception as e:
                print(f"  [WARN] Error inesperado {celda_origen} -> {celda_destino}: {e}")

    wb_readonly.close()

    try:
        mes_nuevo = ws_modelo['B5'].value or ""
        titulo_actual = ws_modelo.title
        anio = titulo_actual[-2:] if titulo_actual[-2:].isdigit() else "26"
        ws_modelo.title = f"MODELO DEUDA {str(mes_nuevo).upper()} {anio}"
        print(f"  [OK] Hoja renombrada a: {ws_modelo.title}")
    except Exception as e:
        print(f"  [WARN] No se pudo renombrar la hoja: {e}")

    print("  [OK] Rotacion completada")

def procesar_y_actualizar_focus(
    archivo_focus: Optional[Union[str, Path]] = None,
    archivo_balance: Optional[Union[str, Path]] = None,
    archivo_situacion: Optional[Union[str, Path]] = None,
    archivo_modelo: Optional[Union[str, Path]] = None,
    archivo_acumulado: Optional[Union[str, Path]] = None,
    output_path: Optional[Union[str, Path]] = None
    ) -> str:
    """
    Función principal que procesa y actualiza el archivo FOCUS.
 
    Args:
        archivo_focus: Ruta al archivo FOCUS base (opcional, se busca automáticamente)
        archivo_balance: Ruta al archivo de balance (opcional)
        archivo_situacion: Ruta al archivo de situación (opcional)
        archivo_modelo: Ruta al archivo de modelo deuda (opcional)
        archivo_acumulado: Ruta al archivo acumulado (opcional)
        output_path: Ruta de salida personalizada (opcional)
 
    Returns:
        str: Ruta al archivo FOCUS actualizado generado
    """
    print("\n=== PROCESADOR FOCUS ===")
    print(f"Directorio base: {BASE_DIR}")
    print(f"Directorio de salidas: {SALIDAS_DIR}")
 
    # Crear directorios necesarios
    SALIDAS_DIR.mkdir(exist_ok=True, parents=True)
    BACKUP_DIR.mkdir(exist_ok=True, parents=True)
 
    # Buscar archivos si no se proporcionaron
    directorio_busqueda = Path('.')
 
    if not archivo_focus:
        print("\nBuscando archivo FOCUS...")
        archivo_focus_path = buscar_archivo(directorio_busqueda, ['FOCUS', 'focus'], ['.xlsx', '.xls'])
        if not archivo_focus_path:
            raise FileNotFoundError("No se encontró el archivo FOCUS")
        archivo_focus = archivo_focus_path
        print(f"  Encontrado: {archivo_focus.name}")
    else:
        archivo_focus = Path(archivo_focus)
 
    # Balance eliminado del flujo — ya no se usa
    archivo_balance = None
 
    if not archivo_situacion:
        print("\nBuscando archivo de Situación...")
        archivo_situacion_path = buscar_archivo(directorio_busqueda, ['situacion', 'situación'], ['.xlsx', '.xls'])
        if archivo_situacion_path:
            archivo_situacion = archivo_situacion_path
            print(f"  Encontrado: {archivo_situacion.name}")
        else:
            print("  [WARN] No se encontró archivo de Situación")
    else:
        archivo_situacion = Path(archivo_situacion)
        print(f"  Archivo de Situación proporcionado: {archivo_situacion.name}")
 
    if not archivo_modelo:
        print("\nBuscando archivo de Modelo Deuda...")
        archivo_modelo_path = buscar_archivo(directorio_busqueda, ['modelo', 'deuda', 'MODELO_DEUDA'], ['.xlsx', '.xls'])
        if archivo_modelo_path:
            archivo_modelo = archivo_modelo_path
            print(f"  Encontrado: {archivo_modelo.name}")
    else:
        archivo_modelo = Path(archivo_modelo)
 
    # Cargar archivo FOCUS
    print(f"\nCargando archivo FOCUS: {archivo_focus}")
    try:
        wb = load_workbook(archivo_focus, data_only=False)
 
        ws: Optional[Worksheet] = None
 
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().upper() == "S22":
                ws = wb[sheet_name]
                print(f"  Hoja S22 encontrada: {sheet_name}")
                break
 
        if ws is None:
            for sheet_name in wb.sheetnames:
                test_ws = wb[sheet_name]
                if not isinstance(test_ws, Worksheet):
                    continue
                try:
                    _ = test_ws['H7']
                    _ = test_ws['Q15']
                    _ = test_ws['Q16']
                    ws = test_ws
                    print(f"  Hoja FOCUS encontrada: {sheet_name}")
                    break
                except Exception:
                    continue
 
        if ws is None:
            for sheet in wb.worksheets:
                if isinstance(sheet, Worksheet):
                    ws = sheet
                    break
            if ws is None:
                raise ValueError("No se encontró ninguna hoja de cálculo válida en el archivo FOCUS")
            print(f"  Usando primera hoja: {ws.title}")
 
        print(f"  Hoja activa: {ws.title}")
        print(f"  Total de hojas en el archivo: {len(wb.worksheets)}")
        print(f"  Nombres de hojas: {', '.join(wb.sheetnames)}")
    except Exception as e:
        raise Exception(f"Error al cargar archivo FOCUS: {e}")
 
    # ── Buscar hoja MODELO DEUDA para TRM/rotación/totales ──────────────────
    ws_modelo: Optional[Worksheet] = None
    for sheet_name in wb.sheetnames:
        if "MODELO" in sheet_name.strip().upper():
            candidate = wb[sheet_name]
            if isinstance(candidate, Worksheet):
                ws_modelo = candidate
                print(f"  Hoja MODELO encontrada: {sheet_name}")
                break
    if ws_modelo is None:
        ws_modelo = ws
        print("  [WARN] No se encontró hoja MODELO, usando hoja activa para TRM")
 
    # Detectar mes actual del FOCUS
    mes_actual = detectar_mes_archivo(archivo_focus)
    mes_siguiente = obtener_mes_siguiente(mes_actual)
 
    print(f"\nMes actual detectado: {mes_actual}")
    print(f"Mes siguiente: {mes_siguiente}")
 
    insertar_dato_entrada(ws_modelo, 'B5', mes_siguiente, f"Mes siguiente ({mes_siguiente})", forzar_sobrescribir=True)

    # --- NUEVO: renombrar la hoja del modelo con el mes correcto ---
    try:
        titulo_actual = ws_modelo.title
        anio = titulo_actual[-2:] if titulo_actual[-2:].isdigit() else "26"
        nuevo_titulo = f"MODELO DEUDA {str(mes_siguiente).upper()} {anio}"
        if ws_modelo.title != nuevo_titulo:
            ws_modelo.title = nuevo_titulo
            print(f"  [OK] Hoja renombrada a: {ws_modelo.title}")
    except Exception as e:
        print(f"  [WARN] No se pudo renombrar la hoja: {e}")
    # --- FIN NUEVO ---
    
    mes_anterior_nombre = ""
    if mes_siguiente in MESES:
        idx_sig = MESES.index(mes_siguiente)
        mes_anterior_nombre = MESES[(idx_sig - 1) % 12]
        insertar_dato_entrada(ws_modelo, 'H7', f'{mes_siguiente}(Mes)', f"Mes actual H7 = {mes_siguiente}(Mes)", forzar_sobrescribir=True)
        insertar_dato_entrada(ws_modelo, 'H8', f'{mes_anterior_nombre}(Mes)', f"Mes anterior H8 = {mes_anterior_nombre}(Mes)", forzar_sobrescribir=True)
 
    trm_info = obtener_trm_detallada(BASE_DIR)
    if trm_info and trm_info.get('eur'):
        j7_anterior = to_float(ws_modelo['J7'].value)
        if j7_anterior > 0 and ws_modelo['J8'].data_type != 'f':
            ws_modelo['J8'].value = j7_anterior
            print(f"  [TRM] J8 actualizado con TRM anterior: {j7_anterior}")
        insertar_dato_entrada(ws_modelo, 'J7', trm_info['eur'],
                              "TRM EUR cierre mes actual J7", forzar_sobrescribir=True)
        print(f"  [TRM] J7 actualizado: {trm_info['eur']}")
 
        j7_nuevo = to_float(trm_info['eur'])
        j8_val = to_float(ws_modelo['J8'].value)
        if j7_nuevo > 0 and j8_val > 0:
           j10_val = (j7_nuevo + j8_val) / 2.0
           ws_modelo['J10'].value = j10_val
           aplicar_formato_numero(ws_modelo, 'J10')
        print(f"  [TRM] J10 (promedio) = ({j7_nuevo} + {j8_val}) / 2 = {j10_val:.3f}")
 
    # ── ROTACIÓN TEMPRANA DESHABILITADA ──────────────────────────────────────
    # La rotación D22→D14, F22→F14, H22→H14, D29→D23 se hace en el bloque
    # de acumulados, leyendo del archivo base ANTES de cualquier modificación.
    print("\n=== ROTACIÓN: se ejecuta en bloque de acumulados ===")
 
    # Inicializar variables
    total_balance = 0.0
    total_situacion = 0.0
    total_60_plus = 0.0
    total_30 = 0.0
    total_vencido = 0.0
    provision = 0.0
    cobros_sit_vencida = 0.0
    cobros_sit_no_vencida = 0.0
    total_situacion_acum = 0.0
 
    if archivo_situacion and archivo_situacion.exists():
        print("\nProcesando Situación...")
        try:
            df_situacion = leer_excel(archivo_situacion)
            cobros_sit_vencida, cobros_sit_no_vencida, total_situacion, total_situacion_acum = procesar_situacion(df_situacion)
            print(f"  [SITUACION] Cobros mes      : {total_situacion:,.2f}")
            print(f"  [SITUACION] Cobros acumulado: {total_situacion_acum:,.2f}")
        except Exception as e:
            print(f"  [ERROR] Error procesando Situación: {e}")
            import traceback; print(traceback.format_exc())
    else:
        print("  [WARN] Archivo de Situación no encontrado")
 
    ### CAMBIO: lectura de la celda fija I18/I19 para Cobro No Vencida (F15) ###
    cobro_no_vencida_celda_fija = 0.0
    if archivo_situacion and archivo_situacion.exists():
        try:
            cobro_no_vencida_celda_fija = leer_celda_fija_situacion(archivo_situacion)
        except Exception as e:
            print(f"  [ERROR] Error leyendo celda fija I18/I19: {e}")
 
    # ── Modelo Deuda externo: fuente de H22/D22/F22/usd_total/GRAN TOTAL ────
    espana_data = {
        'h22': 0.0, 'd22': 0.0, 'f22': 0.0, 'usd_total': 0.0,
        'saldo_total_gran': 0.0, 'venc_30_gran': 0.0, 'suma_vencidos_gran': 0.0,
    }
    if archivo_modelo and archivo_modelo.exists():
        espana_data = procesar_modelo_como_espana(archivo_modelo)
    else:
        print("  [WARN] Sin archivo Modelo Deuda — espana_data queda en ceros")
 
    if archivo_modelo and archivo_modelo.exists():
        print("\nProcesando Modelo Deuda...")
        try:
            h22_calc         = espana_data['h22']
            d22_calc         = espana_data['d22']
            f22_calc         = espana_data['f22']
            usd_total_espana = espana_data['usd_total']
 
            ### CAMBIO: nuevos valores de GRAN TOTAL (G436 / J436:N436) ###
            saldo_total_gran   = espana_data.get('saldo_total_gran', 0.0)    # G436
            venc_30_gran       = espana_data.get('venc_30_gran', 0.0)        # J436 (solo 30 días)
            suma_vencidos_gran = espana_data.get('suma_vencidos_gran', 0.0)  # SUMA(J436:N436)
 
            trm_usd_val  = to_float(trm_info.get('usd', 0)) if trm_info else 0.0
            valor_43042  = (usd_total_espana * trm_usd_val) / 1000.0
            print(f"  [MODELO] H22={h22_calc:,.3f}, D22={d22_calc:,.3f}, F22={f22_calc:,.3f}")
            print(f"  [MODELO] USD={usd_total_espana:,.3f} × TRM {trm_usd_val} = 43042: {valor_43042:,.3f} miles")
            print(f"  [MODELO] GRAN TOTAL: Saldo(G436)={saldo_total_gran:,.3f}, "
                  f"Venc30(J436)={venc_30_gran:,.3f}, SumaVenc(J436:N436)={suma_vencidos_gran:,.3f}")
 
            # Leer iniciales del archivo base (finales del mes anterior)
            # ANTES de escribir nada — estos son los valores correctos
            _wb_init = load_workbook(archivo_focus, data_only=True)
            _ws_init = None
            for _sn in _wb_init.sheetnames:
                if "MODELO" in _sn.strip().upper():
                    _ws_init = _wb_init[_sn]
                    break
            if _ws_init is None:
                _ws_init = _wb_init.active

            d14_val = to_float(_ws_init['D22'].value) or 0.0  # Final vencida → inicial
            f14_val = to_float(_ws_init['F22'].value) or 0.0  # Final no vencida → inicial
            h14_val = to_float(_ws_init['H22'].value) or 0.0  # Final total → inicial
            d23_val = to_float(_ws_init['D29'].value) or 0.0  # Final prov. acum → inicial
            _wb_init.close()

            print(f"  [MODELO] D14(=D22_ant)={d14_val:,.2f}, F14(=F22_ant)={f14_val:,.2f}, H14(=H22_ant)={h14_val:,.2f}")
            print(f"  [MODELO] D23(=D29_ant)={d23_val:,.2f} (provisión acumulada inicial)")

            # Escribir iniciales correctos en el modelo
            if ws_modelo['D14'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'D14', d14_val, f"D14 inicial=D22 ant ({d14_val:,.3f})", forzar_sobrescribir=True)
            if ws_modelo['F14'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'F14', f14_val, f"F14 inicial=F22 ant ({f14_val:,.3f})", forzar_sobrescribir=True)
            if ws_modelo['H14'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'H14', h14_val, f"H14 inicial=H22 ant ({h14_val:,.3f})", forzar_sobrescribir=True)
            ws_modelo['D23'].value = d23_val
            aplicar_formato_numero(ws_modelo, 'D23')
            ws_modelo['H23'].value = d23_val
            aplicar_formato_numero(ws_modelo, 'H23')
            print(f"  [FIX-D23/H23] D23 = H23 = {d23_val:,.3f}")
             
            ### CAMBIO 1: D15 (Cobro vencida) ###
            # Antes: cobros_vencida_mes = cobros_sit_vencida (de Situación)
            # Ahora: SUMA(J436:N436) del Modelo Deuda ÷ 1000, menos D14, en negativo
            cobros_vencida_mes = -abs((suma_vencidos_gran / 1000.0) - d14_val)
            print(f"  [D15] SUMA(J436:N436)/1000 ({suma_vencidos_gran/1000.0:,.3f}) - D14 ({d14_val:,.3f}) "
                  f"= {cobros_vencida_mes:,.3f}")
 
            ### CAMBIO 2: F15 (Cobro no vencida) ###
            # Antes: cobros_no_vencida_mes = cobros_sit_no_vencida (detección dinámica)
            # Ahora: celda fija I18/I19 de Situación ÷ 1000, menos D15 (ya calculado), en negativo
            valor_i18_i19_miles = cobro_no_vencida_celda_fija / 1000.0
            # F15 = H15 del FOCUS anterior - D15
            _wb_h15 = load_workbook(archivo_focus, data_only=True)
            _ws_h15 = None
            for _sn in _wb_h15.sheetnames:
                if "MODELO" in _sn.strip().upper():
                    _ws_h15 = _wb_h15[_sn]
                    break
            if _ws_h15 is None:
                _ws_h15 = _wb_h15.active
            
            _wb_h15.close()

            # Reutilizamos el valor ya calculado dinámicamente más arriba
            # (cobro_no_vencida_celda_fija = leer_celda_fija_situacion(archivo_situacion))
            situacion_valor = cobro_no_vencida_celda_fija
            
            # F15 = -((H20 o H21 / 1000) + D15)
            # CORRECTO (suma con D15 que ya es negativo):
            cobros_no_vencida_mes = -((situacion_valor / 1000) + cobros_vencida_mes)
            print(f"  [F15] -((Situacion({situacion_valor:,.0f})/1000) + D15({cobros_vencida_mes:,.3f})) = {cobros_no_vencida_mes:,.3f}")
            
            total_cobros_mes = cobros_vencida_mes + cobros_no_vencida_mes
 
            insertar_dato_entrada(ws_modelo, 'D15', cobros_vencida_mes,
                                  "Cobros del mes Vencida D15", forzar_sobrescribir=True)
            insertar_dato_entrada(ws_modelo, 'F15', cobros_no_vencida_mes,
                                  "Cobros del mes No Vencida F15", forzar_sobrescribir=True)
 
            h15_val = total_cobros_mes
            insertar_formula(ws_modelo, 'H15', '=+D15+F15',
                             "Cobros totales mes H15 (formula = D15+F15)")
            print(f"  [MODELO] H15(formula)={h15_val:,.2f}  (D15={cobros_vencida_mes:,.2f}, F15={cobros_no_vencida_mes:,.2f})")
 
            ### CAMBIO 3: D17 (Vencido del mes) ###
            # Antes: d17_calc = d22_calc - d14_val - cobros_vencida_mes  (plug)
            # Ahora: lectura directa de "vencido 30 días" del Modelo Deuda ÷ 1000, en positivo
            d17_calc = abs(venc_30_gran / 1000.0)
            insertar_dato_entrada(ws_modelo, 'D17', d17_calc,
                                  f"Vencidos del mes D17 (lectura directa venc.30) = {d17_calc:,.2f}",
                                  forzar_sobrescribir=True)
            insertar_formula(ws_modelo, 'F17', '=-D17',
                             "Vencidos del mes F17 (formula = -D17)")
            f17_calc = -d17_calc
            print(f"  [D17] Venc.30 ({venc_30_gran:,.3f}) / 1000 = {d17_calc:,.3f} (positivo)")
            print(f"  [MODELO] D17={d17_calc:,.2f}, F17(formula)={f17_calc:,.2f}")
 
           ### CAMBIO 4: F16 (Facturación del mes - No Vencida) ###
            saldo_total_gran_miles = saldo_total_gran / 1000.0

            # H22 final = H14 + H15 + H16 + H17... pero F16 depende de H22 → circular
            # En cambio calculamos H22 final desde sus componentes conocidos:
            # H22 = H14 (inicial total) + H15 (cobros) + H17 (vencidos)
            # F16 es la incógnita que despejamos:
            # saldo_total_gran/1000 = H22_final = H14 + H15 + H16 + H17
            # → F16 = H16 = saldo_total_gran/1000 - H14 - H15 - H17

            h14_total = to_float(ws_modelo['H14'].value)  # inicial total post-rotación
            h15_total = cobros_vencida_mes + cobros_no_vencida_mes  # D15+F15
            h17_total = d17_calc + f17_calc  # D17+F17 = d17_calc + (-d17_calc) = 0
            # Pero H17 en el modelo es D17+F17 = d17_calc - d17_calc = 0
            # así que simplificamos:
            h22_final_calc = saldo_total_gran_miles  # G_total/1000 = deuda bruta final
            f16_calc = h22_final_calc - h14_total - h15_total - (d17_calc + f17_calc)

            print(f"  [F16] saldo_total/1000={saldo_total_gran_miles:,.3f}")
            print(f"  [F16] H14={h14_total:,.3f}, H15={h15_total:,.3f}, D17+F17={d17_calc+f17_calc:,.3f}")
            print(f"  [F16] F16 = {saldo_total_gran_miles:,.3f} - {h14_total:,.3f} "
                  f"- {h15_total:,.3f} - {d17_calc+f17_calc:,.3f} = {f16_calc:,.3f}")

            insertar_dato_entrada(ws_modelo, 'F16', f16_calc,
                                  f"Facturación mes F16 = {f16_calc:,.3f}",
                                  forzar_sobrescribir=True)
            insertar_formula(ws_modelo, 'H16', '=F16',
                             "Facturación mes H16 (fórmula = F16)")

            # Limpiar D16 (facturación vencida — no aplica)
            if ws_modelo['D16'].data_type != 'f' and ws_modelo['D16'].value not in (None, 0, 0.0, ''):
                ws_modelo['D16'].value = None
                print("  [MODELO] D16 limpiado")

            print(f"  [MODELO] F16={f16_calc:,.3f}, H16=fórmula(=F16), D16=vacío")
            print("  [MODELO] D22/F22/H22 se calculan vía fórmula (no se sobrescriben)")
 
            ### CAMBIO 5: D24 (Dotación del mes) ###
            ### D24 (Dotación del mes - Vencida) ###
            d24_val = 0.0
            try:
                incobrable_total = procesar_modelo_vencimiento(archivo_modelo)
                incobrable_miles = incobrable_total / 1000.0
                d24_val = round(-incobrable_miles - d23_val, 3)
                print(f"  [D24] -(incobrable/1000) - D23")
                print(f"  [D24] -({incobrable_miles:,.3f}) - ({d23_val:,.3f}) = {d24_val:,.3f}")
            except Exception as e:
                print(f"  [WARN] No se pudo calcular D24: {e}")
            
            if d24_val != 0:
                insertar_dato_entrada(ws_modelo, 'D24', d24_val,
                                      "Dotaciones del mes D24", forzar_sobrescribir=True)
                print(f"  [MODELO] D24 = {d24_val:,.3f}")
            else:
                print("  [WARN] D24 no calculado (valor = 0)")
 
            # ── Acumulados ──────────────────────────────────────────────────
            # ── Acumulados ──────────────────────────────────────────────────
            print("\n  === LEYENDO ACUMULADOS DEL FOCUS BASE (MES ANTERIOR) ===")
            try:
                wb_base = load_workbook(archivo_focus, data_only=True)
                ws_base_modelo = None
                for sn in wb_base.sheetnames:
                    if "MODELO" in sn.strip().upper():
                        ws_base_modelo = wb_base[sn]
                        break
                if ws_base_modelo is None:
                    ws_base_modelo = wb_base.active

                def leer_base(celda):
                    return to_float(ws_base_modelo[celda].value)

                # Acumulados del mes anterior
                d36_anterior = leer_base('D36')
                f36_anterior = leer_base('F36')
                d37_anterior = leer_base('D37')
                f37_anterior = leer_base('F37')
                d38_anterior = leer_base('D38')
                f38_anterior = leer_base('F38')
                d45_anterior = leer_base('D45')
                f42_anterior = leer_base('F42')
                d46_anterior = leer_base('D46')
                f51_anterior = leer_base('F51')

                # Valores del mes actual para dotaciones
                f21_val = to_float(ws_modelo['F21'].value)

                # Desdotaciones del mes actual: dato de entrada SOLO de este mes.
                # Si no hay desdotación nueva este mes, debe quedar en 0.
                d25_val = to_float(ws_modelo['D25'].value)
                f30_val = to_float(ws_modelo['F30'].value)

                wb_base.close()

                print(f"  D36_ant={d36_anterior:,.3f}  F36_ant={f36_anterior:,.3f}")
                print(f"  D37_ant={d37_anterior:,.3f}  F37_ant={f37_anterior:,.3f}")
                print(f"  D38_ant={d38_anterior:,.3f}  F38_ant={f38_anterior:,.3f}")
                print(f"  D45_ant={d45_anterior:,.3f}  F42_ant={f42_anterior:,.3f}")
                print(f"  D46_ant={d46_anterior:,.3f}  F51_ant={f51_anterior:,.3f}")
                print(f"  D25_mes={d25_val:,.3f}  F30_mes={f30_val:,.3f}")

            except Exception as e:
                print(f"  [WARN] No se pudieron leer acumulados del FOCUS base: {e}")
                d36_anterior = f36_anterior = 0.0
                f37_anterior = d37_anterior = 0.0
                d38_anterior = f38_anterior = 0.0
                d45_anterior = f42_anterior = 0.0
                d46_anterior = f51_anterior = 0.0
                f21_val = d25_val = f30_val = 0.0

            d16_val = to_float(ws_modelo['D16'].value) or 0.0

            # Cobros: D36 = D36_ant + D15 | F36 = F36_ant + F15
            if ws_modelo['D36'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'D36', d36_anterior + cobros_vencida_mes,
                                      f"Cobros acum D36 = {d36_anterior:,.3f} + {cobros_vencida_mes:,.3f}",
                                      forzar_sobrescribir=True)
            if ws_modelo['F36'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'F36', f36_anterior + cobros_no_vencida_mes,
                                      f"Cobros acum F36 = {f36_anterior:,.3f} + {cobros_no_vencida_mes:,.3f}",
                                      forzar_sobrescribir=True)

            # Facturación: D37 = D37_ant + D16 | F37 = F37_ant + F16
            if ws_modelo['D37'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'D37', d37_anterior + d16_val,
                                      f"Facturación acum D37 = {d37_anterior:,.3f} + {d16_val:,.3f}",
                                      forzar_sobrescribir=True)
            if ws_modelo['F37'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'F37', f37_anterior + f16_calc,
                                      f"Facturación acum F37 = {f37_anterior:,.3f} + {f16_calc:,.3f}",
                                      forzar_sobrescribir=True)

            # +/- Vencidos: D38 = D38_ant + D17 | F38 = F38_ant - D17
            if ws_modelo['D38'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'D38', d38_anterior + d17_calc,
                                      f"Vencidos acum D38 = {d38_anterior:,.3f} + {d17_calc:,.3f}",
                                      forzar_sobrescribir=True)
            if ws_modelo['F38'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'F38', f38_anterior - d17_calc,
                                      f"Vencidos acum F38 = {f38_anterior:,.3f} - {d17_calc:,.3f}",
                                      forzar_sobrescribir=True)

            # Dotaciones: D45 = D45_ant + D24 | F42 = F42_ant + F21
            if ws_modelo['D45'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'D45', d45_anterior + d24_val,
                                      f"Dotaciones acum D45 = {d45_anterior:,.3f} + {d24_val:,.3f}",
                                      forzar_sobrescribir=True)
            if ws_modelo['F42'].data_type != 'f':
                insertar_dato_entrada(ws_modelo, 'F42', f42_anterior + f21_val,
                                      f"Dotaciones acum F42 = {f42_anterior:,.3f} + {f21_val:,.3f}",
                                      forzar_sobrescribir=True)

            # Desdotaciones: D46 = D46_ant + D25 (sin abs(), respeta el signo real)
            d46_nuevo = d46_anterior + d25_val
            insertar_dato_entrada(ws_modelo, 'D46', d46_nuevo,
                                  f"Desdotaciones acum D46 = {d46_anterior:,.3f} + {d25_val:,.3f} = {d46_nuevo:,.3f}",
                                  forzar_sobrescribir=True)

            # IMPORTANTE: limpiar D25 después de sumarlo a D46.
            # D25 es un dato de entrada solo del mes en curso; si no se
            # reinicia a 0, el mes siguiente este mismo valor se vuelve a
            # sumar a D46 (arrastre/doble conteo indebido).
            # D25 NUNCA debe ser fórmula (es un dato manual), así que se
            # limpia siempre, sin condición de por medio.
            valor_d25_previo = ws_modelo['D25'].value
            ws_modelo['D25'].value = 0
            print(f"  [CLEAN] D25 reiniciado a 0 (valor anterior era: {valor_d25_previo}) "
                  f"después de acumular {d25_val:,.3f} en D46")
            
            insertar_formula(ws_modelo, 'F51', '=H43+50',
                 "Desdotaciones acum F51 (fórmula = H43+50)")

            print(f"  [MODELO] Acumulados actualizados: D36/F36, D37/F37, D38/F38, D45/F42, D46/F51")

        except Exception as e:
            print(f"  [ERROR] Error procesando archivo de Modelo Deuda: {str(e)}")
            import traceback
            print(f"  [DEBUG] {traceback.format_exc()}")
    else:
        print("  [WARN] Archivo de Modelo Deuda no encontrado o no existe")

    print("  [INFO] Q22: no se actualiza (balance eliminado)")
 
    # ── Llamar actualizar_hoja_focus ─────────────────────────────────────────
    trm_usd_val = espana_data.get('trm_usd_archivo', 0.0)
    if trm_usd_val == 0.0:
     trm_usd_val = to_float(trm_info.get('usd', 0)) if trm_info else 0.0
    espana_data_final = espana_data if 'espana_data' in locals() else {
        'h22': 0.0, 'd22': 0.0, 'f22': 0.0, 'usd_total': 0.0,
        'saldo_total_gran': 0.0, 'venc_30_gran': 0.0, 'suma_vencidos_gran': 0.0,
    }
 
    actualizar_hoja_focus(
        wb=wb,
        mes_siguiente=mes_siguiente,
        trm_usd_val=trm_usd_val,
        espana_data=espana_data_final,
    )
 
    # ── Actualizar TRM en hoja FOCUS ─────────────────────────────────────────
    actualizar_celdas_trm(ws, BASE_DIR, mes_actual)
 
    # ── Validar totales ───────────────────────────────────────────────────────
    if total_vencido > 0:
        wb_values = load_workbook(archivo_focus, data_only=True)
        ws_values = None
        for sn in wb_values.sheetnames:
            if sn.strip().upper() == "S22" or "MODELO" in sn.strip().upper():
                ws_values = wb_values[sn]
                break
        validar_datos(ws_modelo, ws_values, total_vencido * 1000, total_vencido)
        wb_values.close()
 
    # ── Definir output_path y crear backup ───────────────────────────────────
    if not output_path:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = SALIDAS_DIR / f"FOCUS_ACTUALIZADO_{timestamp}.xlsx"
    else:
        output_path = Path(output_path)
 
    backup_path = BACKUP_DIR / f"{archivo_focus.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{archivo_focus.suffix}"
    shutil.copy2(archivo_focus, backup_path)
    print(f"\nBackup creado: {backup_path}")
 
    # ── Guardar archivo final ─────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\n[OK] Archivo FOCUS actualizado guardado: {output_path}")
    print(f"[OK] Tamaño: {output_path.stat().st_size / 1024:.2f} KB")
    return str(output_path)
 
def actualizar_hoja_focus(
    wb: openpyxl.Workbook,
    mes_siguiente: str,
    trm_usd_val: float,
    espana_data: dict,
) -> bool:
    print("\n=== ACTUALIZANDO HOJA FOCUS ===")

    ws_focus: Optional[Worksheet] = None
    for sn in wb.sheetnames:
        sn_upper = sn.strip().upper()
        if sn_upper == "FOCUS" or sn_upper.startswith("FOCUS"):
            ws_focus = wb[sn]
            print(f"  Hoja FOCUS encontrada: '{sn}'")
            break
    if ws_focus is None:
        # Último recurso: imprimir hojas disponibles para diagnóstico
        print(f"  [WARN] No se encontró hoja 'FOCUS'. Hojas disponibles: {wb.sheetnames}")
        return False
    print(f"  Hoja FOCUS: '{ws_focus.title}'")

    FILA_ENCABEZADO_MES = 6   # fila con Marzo_ac, Febrero_ac, etc.
    FILA_DATOS_INICIO   = 8   # primera fila de datos
    FILA_DATOS_FIN      = 15  # última fila de datos (BCA10_1)
    COL_CODIGO          = 1   # col A

    # Dos bloques fijos verificados contra el archivo real
    BLOQUES = [
        {"nombre": "MONEDA LOCAL", "col_inicio": 3,  "col_fin": 7},   # C-G
        {"nombre": "Euros",        "col_inicio": 9,  "col_fin": 13},  # I-M
    ]

    for bloque in BLOQUES:
        col_ini = bloque["col_inicio"]
        col_fin = bloque["col_fin"]
        nombre  = bloque["nombre"]
        print(f"\n  Shift bloque {nombre} (cols {get_column_letter(col_ini)}-{get_column_letter(col_fin)})...")

        # Shift de derecha a izquierda: col_fin <- col_fin-1 <- ... <- col_ini+1 <- col_ini
        for col_dest in range(col_fin, col_ini, -1):
            col_src = col_dest - 1

            # Fila 6 (encabezado de mes)
            ws_focus.cell(row=FILA_ENCABEZADO_MES, column=col_dest).value = \
                ws_focus.cell(row=FILA_ENCABEZADO_MES, column=col_src).value

            # Filas de datos
            for row in range(FILA_DATOS_INICIO, FILA_DATOS_FIN + 1):
                cell_src  = ws_focus.cell(row=row, column=col_src)
                cell_dest = ws_focus.cell(row=row, column=col_dest)
                if cell_src.data_type == 'f' and cell_src.value:
                    cell_dest.value = _ajustar_formula_columna(
                        str(cell_src.value), col_src, col_dest
                    )
                else:
                    cell_dest.value = cell_src.value
                if cell_src.number_format and cell_src.number_format != 'General':
                    cell_dest.number_format = cell_src.number_format

        # Escribir mes nuevo en col_ini fila 6
        ws_focus.cell(row=FILA_ENCABEZADO_MES, column=col_ini).value = f"{mes_siguiente}_ac"
        print(f"    {get_column_letter(col_ini)}6 = '{mes_siguiente}_ac'")

        # Limpiar col_ini filas de datos (col nueva vacía inicialmente)
        for row in range(FILA_DATOS_INICIO, FILA_DATOS_FIN + 1):
            ws_focus.cell(row=row, column=col_ini).value = None

    # Fix BCA10_1: corregir SUM(X8:X14) -> SUM(X9:X14) en cols históricas
    fila_bca = FILA_DATOS_FIN  # fila 15
    fixes = 0
    for bloque in BLOQUES:
        for col_idx in range(bloque["col_inicio"] + 1, bloque["col_fin"] + 1):  # D en adelante (no C, ya limpiada)
            cell = ws_focus.cell(row=fila_bca, column=col_idx)
            if cell.data_type == 'f' and cell.value:
                nuevo, n = _fix_bca10_sum(str(cell.value), col_idx)
                if n > 0:
                    cell.value = nuevo
                    fixes += 1
    print(f"\n  Fix BCA10_1: {fixes} celdas corregidas")

    # Escribir 43001 (col C = col_inicio bloque MONEDA LOCAL = 3)
    h22_val = espana_data.get('h22', 0.0)
    fila_43001 = _buscar_fila_por_codigo(ws_focus, '43001', FILA_DATOS_INICIO, FILA_DATOS_FIN, COL_CODIGO)
    if fila_43001:
        ws_focus.cell(row=fila_43001, column=3).value = h22_val
        aplicar_formato_numero(ws_focus, f"C{fila_43001}")
        print(f"  43001 C{fila_43001} = {h22_val:,.3f}")
    else:
        print("  [WARN] No se encontró fila 43001")

    # Calcular y escribir 43042
    usd_total   = espana_data.get('usd_total', 0.0)
    valor_43042 = (usd_total * trm_usd_val) / 1000.0
    fila_43042  = _buscar_fila_por_codigo(ws_focus, '43042', FILA_DATOS_INICIO, FILA_DATOS_FIN, COL_CODIGO)
    if fila_43042:
        ws_focus.cell(row=fila_43042, column=3).value = valor_43042
        aplicar_formato_numero(ws_focus, f"C{fila_43042}")
        print(f"  43042 C{fila_43042} = {usd_total:,.3f} USD × {trm_usd_val} / 1000 = {valor_43042:,.3f}")
    else:
        print("  [WARN] No se encontró fila 43042")

    # Recalcular BCA10_1 col C = SUM(C9:C14) numérico
    suma = sum(to_float(ws_focus.cell(row=r, column=3).value) for r in range(9, 15))
    ws_focus.cell(row=fila_bca, column=3).value = suma
    aplicar_formato_numero(ws_focus, f"C{fila_bca}")
    print(f"  BCA10_1 C{fila_bca} = SUM(C9:C14) = {suma:,.3f}")

    print("  [OK] actualizar_hoja_focus completado")
    return True
 
 
# ── Helpers privados ─────────────────────────────────────────────────────────
 
def _buscar_fila_por_codigo(
    ws: Worksheet,
    codigo: str,
    fila_inicio: int,
    fila_fin: int,
    col_codigo: int,
) -> Optional[int]:
    """Devuelve el número de fila donde col_codigo contiene 'codigo' (búsqueda parcial)."""
    codigo_norm = codigo.strip().upper()
    for row_idx in range(fila_inicio, fila_fin + 1):
        val = str(ws.cell(row=row_idx, column=col_codigo).value or "").strip().upper()
        # Coincidencia exacta primero
        if val == codigo_norm:
            return row_idx
        # Coincidencia parcial (ej. "TOTAL CUENTA OBJETO 43001")
        if codigo_norm in val:
            return row_idx
    return None
 
 
def _ajustar_formula_columna(formula: str, col_src: int, col_dest: int) -> str:
    """
    Desplaza las referencias de columna dentro de una fórmula de col_src a col_dest.
    Solo ajusta referencias absolutas de columna (letras) que coincidan con col_src.
    Preserva referencias con $ y referencias a otras columnas intactas.
 
    Ej: col_src=3 (C), col_dest=4 (D):
        =SUM(C9:C14)  ->  =SUM(D9:D14)
        =C8           ->  =D8
        =$A8          ->  =$A8   (col absoluta A no cambia)
    """
    import re as _re
    letra_src  = openpyxl.utils.get_column_letter(col_src)
    letra_dest = openpyxl.utils.get_column_letter(col_dest)
 
    # Reemplazar referencias a la columna origen (sin $ delante de la letra)
    # Patrón: no precedido por $ ni letra, seguido de dígito o :
    patron = _re.compile(
        rf'(?<![A-Z$]){_re.escape(letra_src)}(?=\d|:)',
        _re.IGNORECASE
    )
    return patron.sub(letra_dest, formula)
 
 
def _fix_bca10_sum(formula: str, col_idx: int) -> tuple:
    """
    Si la fórmula es =SUM(Xn:X14) con n < 9, la corrige a =SUM(X9:X14).
    Retorna (formula_corregida, numero_reemplazos).
    """
    import re as _re
    col_ltr = openpyxl.utils.get_column_letter(col_idx)
 
    # Patrón: =SUM(<COL><FILA_INICIO>:<COL>14)  con fila_inicio < 9
    patron = _re.compile(
        rf'=SUM\({_re.escape(col_ltr)}([1-8]):{_re.escape(col_ltr)}14\)',
        _re.IGNORECASE
    )
    nueva_formula, n = patron.subn(
        lambda m: f'=SUM({col_ltr}9:{col_ltr}14)',
        formula
    )
    return nueva_formula, n

if __name__ == "__main__":
    try:
        cli_args = sys.argv[1:]
        situacion_arg = cli_args[0] if len(cli_args) > 0 else None
        focus_arg     = cli_args[1] if len(cli_args) > 1 else None
        modelo_arg    = cli_args[2] if len(cli_args) > 2 else None

        if cli_args:
            print("[CLI] Ejecutando con argumentos:")
            print(f"  situacion: {situacion_arg}")
            print(f"  focus:     {focus_arg}")
            print(f"  modelo:    {modelo_arg}")

        resultado = procesar_y_actualizar_focus(
            archivo_focus=focus_arg,
            archivo_balance=None,
            archivo_situacion=situacion_arg,
            archivo_modelo=modelo_arg,
            archivo_acumulado=None,
        )
        print(f"[OK] Archivo generado: {resultado}")
    except FileNotFoundError as e:
        logger.error(f"Error de archivo no encontrado: {e}")
        print(f"\n[ERROR] ERROR: {e}")
        raise
    except pd.errors.EmptyDataError as e:
        logger.error(f"Archivo vacío o inválido: {e}")
        raise ValueError(f"Archivo vacío o inválido: {e}")
    except Exception as e:
        logger.error(f"Error inesperado: {e}")
        logger.error(traceback.format_exc())
        print(f"\n[ERROR] ERROR INESPERADO: {e}")
        raise